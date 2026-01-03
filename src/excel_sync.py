
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Path to the Excel file
EXCEL_FILE = "Pharmacy_Ledger.xlsx"

def sync_to_excel(data: dict):
    """
    Syncs a single transaction record to a local Excel file.
    :param data: Dictionary containing transaction details.
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            # Create a new workbook and verify headers
            wb = Workbook()
            ws = wb.active
            ws.title = "Pharmacy Ledger"
            headers = [
                "Date", "Total Sale", "Cash Sale", "Card", "Talabat", 
                "Insurance", "Credit Sale", "Med Purchase", "Other Exp", 
                "Collection (Owner)", "Cash In Hand", "Discrepancy"
            ]
            ws.append(headers)
        else:
            # Load existing workbook
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        # Prepare row data
        row = [
            data.get('date', ''),
            data.get('total_sale', 0),
            data.get('cash_sale', 0),
            data.get('card_sale', 0),
            data.get('talabat_sale', 0),
            data.get('insurance_sale', 0),
            data.get('credit_sale', 0),
            data.get('med_purchase', 0),
            data.get('other_exp', 0),
            data.get('owner_collection', 0),
            data.get('closing_petty', 0),
            data.get('discrepancy', 0)
        ]

        # Append row
        ws.append(row)
        
        # Save file
        wb.save(EXCEL_FILE)
        print(f"Successfully synced entry to {EXCEL_FILE}.")

    except Exception as e:
        print(f"Failed to sync to Excel: {e}")
